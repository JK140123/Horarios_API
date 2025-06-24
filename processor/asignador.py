import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import re
import unicodedata
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import io

# --- Parámetros ---
salones = [
    "ESCENARIO SIMULADO 1 ID:143", "ESCENARIO SIMULADO 2 ID:144", "ESCENARIO SIMULADO 3 ID:145",
    "ESCENARIO SIMULADO 4 ID:146", "ESCENARIO SIMULADO 5 ID:147", "ESCENARIO SIMULADO 6 ID:148",
    "ESCENARIO SIMULADO 7 ID:150", "ESCENARIO CX-TPR ID:149", "URGENCIAS 1 ID:153", "URGENCIAS 2 ID:154",
    "HOSPITALIZACION 1 ID:151", "HOSPITALIZACION 2 ID:152", "G-108-CAMP", "G-109-CAMP",
    "G-115 NEUROREHABILITACIÓN ID:G115-CAMP", "G -116 ELECTROFISIOLOGÍA ID:116-CAMP",
    "G-114/ESC. EXAMEN FISICO E INTER. N. 1  ID:G117-CAMP",
    "SALA DE OBSERVACION N. 8 - DESARROLLO, INNOVACION Y PROTOTIPADO  ID:0235",
    "ESC. EXAMEN FISICO E INTER. N.2 ID:0236", "LABORATORIO DE MOVIMIENTO  ID:0237",
    "ESCENARIO SIMULADO 8 ID:0238"
]
capacidad_salones = [10, 10, 10, 10, 10, 10, 12, 12, 10, 10, 10, 10, 30, 30, 20, 30, 20, 20, 20, 20, 20]
dias_semana = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado']

colores_programa = {
    "medicina": "FFFF00",             # Amarillo
    "enfermería": "C9DAF8",           # Azul claro
    "fisioterapia": "EAD1DC",         # Rosa pálido
    "psicología": "9370DB",           # Lila claro
    "educación continua": "93C47D",   # Verde
    "otros": "93C47D"                 # Verde (mismo que educación continua)
}


def normalizar(texto):
    if not isinstance(texto, str):
        return ""
    texto = texto.lower().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    return texto

def obtener_color(programa):
    programa = normalizar(programa)
    if "enfermeria" in programa:
        return PatternFill(start_color=colores_programa["enfermería"], fill_type="solid")
    elif "fisioterapia" in programa:
        return PatternFill(start_color=colores_programa["fisioterapia"], fill_type="solid")
    elif "psicologia" in programa:
        return PatternFill(start_color=colores_programa["psicología"], fill_type="solid")
    elif "educacion continua" in programa or "educacion" in programa:
        return PatternFill(start_color=colores_programa["educación continua"], fill_type="solid")
    elif "medicina" in programa:
        return PatternFill(start_color=colores_programa["medicina"], fill_type="solid")
    elif "otros" in programa:
        return PatternFill(start_color=colores_programa["otros"], fill_type="solid")
    return None

def limpiar_nombre_hoja(nombre):
    return re.sub(r'[\\/*?:\[\]]', '-', nombre)[:31]

def procesar_horarios(df_input):
    df = df_input.copy()
    df.columns = df.columns.str.strip()
    
    def generar_plantilla_con_fechas(fechas_lunes):
        horas = list(range(6, 19))
        columnas = [(d, (lunes + timedelta(days=dias_semana.index(d))).strftime('%d/%m')) for lunes in fechas_lunes for d in dias_semana]
        return pd.DataFrame(index=horas, columns=pd.MultiIndex.from_tuples(columnas, names=["Día", "Fecha"]))

    def obtener_fechas_recurrentes(inicio, fin, dia_semana):
        dia_num = dias_semana.index(dia_semana)
        fechas = []
        actual = inicio + timedelta((dia_num - inicio.weekday()) % 7)
        while actual <= fin:
            fechas.append(actual)
            actual += timedelta(weeks=1)
        return fechas

    def encontrar_salon(programa, escenario_especifico, necesita_grande, necesita_gesell, estudiantes, fechas, dia, hora_ini, hora_fin, cronogramas):

        duracion = hora_fin - hora_ini
    
        def salon_disponible(salon):

            for fecha in fechas:

                columna = (dia, fecha.strftime('%d/%m'))

                for hora in range(hora_ini, hora_fin):

                    if hora in cronogramas[salon].index and columna in cronogramas[salon].columns:

                        if pd.notna(cronogramas[salon].at[hora, columna]):

                            return False

            return True
    
        # 1. Escenario específico (sin validar capacidad)

        if escenario_especifico and escenario_especifico in salones:

            if salon_disponible(escenario_especifico):

                return escenario_especifico
    
        # 2. Necesita cámara Gesell (sin validar capacidad)

        if necesita_gesell and necesita_gesell in salones:

            if salon_disponible(necesita_gesell):

                return necesita_gesell
    
        # 3. Escenario grande (sin validar capacidad)

        if necesita_grande:

            grandes = [

                "ESCENARIO SIMULADO 1 ID:143", "ESCENARIO SIMULADO 2 ID:144",

                "URGENCIAS 1 ID:153", "URGENCIAS 2 ID:154",

                "HOSPITALIZACION 1 ID:151", "HOSPITALIZACION 2 ID:152"

            ]

            for s in grandes:

                if s in salones and salon_disponible(s):

                    return s
    
        # 4. Salón más pequeño disponible (sí valida capacidad)

        salones_disponibles = [

            (i, s) for i, s in enumerate(salones)

            if capacidad_salones[i] >= estudiantes and salon_disponible(s)

        ]

        salones_disponibles.sort(key=lambda x: capacidad_salones[x[0]])
    
        if salones_disponibles:

            return salones_disponibles[0][1]
    
        return None  # Ningún salón disponible

 

    df['Fecha de inicio'] = pd.to_datetime(df['Fecha de inicio'], errors='coerce')
    df['Fecha de fin'] = pd.to_datetime(df['Fecha de fin'], errors='coerce')
    df["Salón asignado"] = None
    df["Salón reasignado"] = None
    df["Hora reasignada"] = None
    df["Fecha fin reasignada"] = None

    inicio_cronograma = df['Fecha de inicio'].min()
    fin_cronograma = df['Fecha de fin'].max()

    fechas_lunes = []
    actual = inicio_cronograma - timedelta(days=inicio_cronograma.weekday())
    while actual <= fin_cronograma:
        fechas_lunes.append(actual)
        actual += timedelta(weeks=1)

    cronogramas = {s: generar_plantilla_con_fechas(fechas_lunes) for s in salones}
    
    # --- Ordenar por prioridad: específico > grande > Gesell > sin requisito ---
    df["__prioridad_especifico"] = df["¿Se necesita un escenario especifico?"].fillna("") == "Sí"
    df["__prioridad_grande"] = df["¿Se necesita un escenario grande?"].fillna("") == "Sí"
    df["__prioridad_gesell"] = df["¿Se necesita camara de Gesell ?"].fillna("") != "No"
    
    def calcular_prioridad(row):
        if row["__prioridad_especifico"]:
            return 0
        elif row["__prioridad_grande"]:
            return 1
        elif row["__prioridad_gesell"]:
            return 2
        else:
            return 3
    
    df["__nivel_prioridad"] = df.apply(calcular_prioridad, axis=1)
    df = df.sort_values(by="__nivel_prioridad").reset_index(drop=True)
    df.drop(columns=["__prioridad_especifico", "__prioridad_grande", "__prioridad_gesell", "__nivel_prioridad"], inplace=True)

    for idx, row in df.iterrows():
        try:
            programa = row.get('Programa')
            asignatura = row.get('Asignatura')
            profesor = row.get('Profesor')
            estudiantes = row.get('Número de estudiantes', 0)
            fecha_ini = row.get('Fecha de inicio')
            fecha_fin = row.get('Fecha de fin')
            dia = row.get('Día de la semana')
            hora_ini = row.get('Hora de inicio')
            hora_fin = row.get('Hora de finalización')
            necesita_escenario = row.get('¿Se necesita un escenario especifico?', '') == "Sí"
            escenario = row.get('Especifica el escenario') if pd.notna(row.get('Especifica el escenario')) else None
            necesita_grande = row.get('¿Se necesita un escenario grande?', '') == "Sí"
            necesita_gesell = row.get("\u00bfSe necesita camara de Gesell ?", "No")

            if pd.isna(hora_ini) or pd.isna(hora_fin) or pd.isna(dia) or pd.isna(fecha_ini):
                continue

            hora_ini_int = int(str(hora_ini).split(":")[0])
            hora_fin_int = int(str(hora_fin).split(":")[0])
            duracion = hora_fin_int - hora_ini_int

            fecha_ini = max(fecha_ini, inicio_cronograma)
            fecha_fin = fin_cronograma if pd.isna(fecha_fin) else min(fecha_fin, fin_cronograma)
            fechas = obtener_fechas_recurrentes(fecha_ini, fecha_fin, dia)
            columnas = [(dia, fecha.strftime('%d/%m')) for fecha in fechas]

            salon_original = encontrar_salon(programa, escenario, necesita_grande, necesita_gesell, estudiantes, fechas, dia, hora_ini_int, hora_fin_int, cronogramas)
            if not salon_original:
                continue

            df.at[idx, "Salón asignado"] = salon_original
            original_disponible = all(
                all(pd.isna(cronogramas[salon_original].at[hora, columna]) for hora in range(hora_ini_int, hora_fin_int))
                for columna in columnas
            )

            if original_disponible:
                for columna in columnas:
                    for hora in range(hora_ini_int, hora_fin_int):
                        cronogramas[salon_original].at[hora, columna] = f"{asignatura} - {profesor}"
            else:
                reubicado = False
                for i, s in enumerate(salones):
                    if s == salon_original or capacidad_salones[i] < estudiantes:
                        continue
                    if all(
                        all(pd.isna(cronogramas[s].at[hora, columna]) for hora in range(hora_ini_int, hora_fin_int))
                        for columna in columnas
                    ):
                        for columna in columnas:
                            for hora in range(hora_ini_int, hora_fin_int):
                                cronogramas[s].at[hora, columna] = f"{asignatura} - {profesor}"
                        df.at[idx, "Salón reasignado"] = s
                        df.at[idx, "Hora reasignada"] = f"{hora_ini_int}:00 - {hora_fin_int}:00"
                        df.at[idx, "Fecha fin reasignada"] = max(fechas)
                        reubicado = True
                        break

                if not reubicado:
                    for nueva_hora_ini in range(6, 19 - duracion + 1):
                        nueva_hora_fin = nueva_hora_ini + duracion
                        if all(
                            all(pd.isna(cronogramas[salon_original].at[h, columna]) for h in range(nueva_hora_ini, nueva_hora_fin))
                            for columna in columnas
                        ):
                            for columna in columnas:
                                for h in range(nueva_hora_ini, nueva_hora_fin):
                                    cronogramas[salon_original].at[h, columna] = f"{asignatura} - {profesor}"
                            df.at[idx, "Hora reasignada"] = f"{nueva_hora_ini}:00 - {nueva_hora_fin}:00"
                            df.at[idx, "Fecha fin reasignada"] = max(fechas)
                            break
        except Exception as e:
            continue

    # --- Crear archivo Excel ---
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Reasignaciones"

    cols_df = ["Programa", "Asignatura", "Profesor", "Fecha de inicio", "Fecha de fin", "Día de la semana",
               "Hora de inicio", "Hora de finalización", "Salón asignado", "Salón reasignado", "Hora reasignada", "Fecha fin reasignada"]
    for r in dataframe_to_rows(df[cols_df], index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for column_cells in ws1.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws1.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    for salon, tabla in cronogramas.items():
        ws = wb.create_sheet(title=limpiar_nombre_hoja(salon))
        ws.append(["Hora"] + [f"{d} {f}" for d, f in tabla.columns])
        for hora in tabla.index:
            fila = [f"{hora}:00"]
            for col in tabla.columns:
                val = tabla.at[hora, col]
                fila.append(val if pd.notna(val) else "")
            ws.append(fila)
        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            contenido_anterior = None
            inicio_fila = None
            for row_idx in range(2, ws.max_row + 2):
                celda = ws.cell(row=row_idx, column=col_idx)
                valor = celda.value
                if valor == contenido_anterior:
                    continue
                if contenido_anterior and inicio_fila:
                    if row_idx - inicio_fila > 1:
                        ws.merge_cells(start_row=inicio_fila, start_column=col_idx, end_row=row_idx - 1, end_column=col_idx)
                    asignatura = contenido_anterior.split(" - ")[0]
                    clase_df = df[df["Asignatura"].str.contains(asignatura, na=False, case=False)]
                    if not clase_df.empty:
                        programa = clase_df.iloc[0]["Programa"]
                        color = obtener_color(programa)
                        for r in range(inicio_fila, row_idx):
                            celda_color = ws.cell(row=r, column=col_idx)
                            if color:
                                celda_color.fill = color
                            celda_color.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                contenido_anterior = valor
                inicio_fila = row_idx
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "B2"
        ws.column_dimensions['A'].width = 8
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream.getvalue()